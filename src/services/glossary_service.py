"""
services/glossary_service.py
Business logic for glossary CRUD operations.
"""

from __future__ import annotations

import io
import logging
import re
import uuid
from typing import Any, Optional

from fastapi import HTTPException
from openpyxl import Workbook, load_workbook  # type: ignore
from openpyxl.styles import Alignment, Font, PatternFill  # type: ignore

from models.schemas import GlossaryEntryCreate, GlossaryEntryUpdate
from repositories.glossary_repository import GlossaryRepository

logger = logging.getLogger(__name__)

_VALID_SCOPES = {"global", "local", "functional"}
_VALID_LANGUAGES = {"de", "es"}
_SHEET_PATTERN = re.compile(r"^EN\s*>\s*(\w{2})\s+terms$", re.IGNORECASE)
_NO_TRANSLATE_PREFIX = re.compile(r"^no\s+translation\s*", re.IGNORECASE)
_PAREN_HINT = re.compile(r"\(([^)]+)\)")
_LANG_LABELS = {"de": "DE", "es": "ES"}


class GlossaryService:
    """Encapsulates glossary business rules on top of the repository layer."""

    def __init__(self, repository: GlossaryRepository) -> None:
        self._repo = repository

    # ── Create ──────────────────────────────────────────────────────

    def create_entry(self, data: GlossaryEntryCreate) -> dict[str, Any]:
        """Create a new glossary entry. Raises 409 on duplicate term+scope."""
        glossary_id = uuid.uuid4().hex

        try:
            self._repo.create(
                glossary_id=glossary_id,
                term=data.term,
                scope=data.scope.value,
                do_not_translate=data.do_not_translate,
                translations=(
                    data.translations.model_dump(exclude_none=True)
                    if data.translations else None
                ),
                comments=data.comments,
                de_comments=data.de_comments,
                es_comments=data.es_comments,
            )
        except Exception as exc:
            if "UNIQUE constraint failed" in str(exc):
                existing = self._repo.get_by_term_and_scope(data.term, data.scope.value)
                raise HTTPException(
                    status_code=409,
                    detail={
                        "message": f"Glossary entry '{data.term}' already exists for scope '{data.scope.value}'",
                        "glossary_id": existing["glossary_id"] if existing else None,
                    },
                )
            raise

        entry = self._repo.get_by_id(glossary_id)
        logger.info("Created glossary entry: glossary_id=%s, term='%s'", glossary_id, data.term)
        return entry  # type: ignore[return-value]

    # ── List ────────────────────────────────────────────────────────

    def list_entries(self, scope: Optional[str] = None) -> dict[str, Any]:
        """Return all active glossary entries, optionally filtered by scope."""
        items = self._repo.list_all(scope=scope)
        return {"items": items, "total": len(items)}

    # ── Get ─────────────────────────────────────────────────────────

    def get_entry(self, glossary_id: str) -> dict[str, Any]:
        """Fetch a single glossary entry or raise 404."""
        entry = self._repo.get_by_id(glossary_id)
        if not entry:
            raise HTTPException(status_code=404, detail="Glossary entry not found")
        return entry

    # ── Update ──────────────────────────────────────────────────────

    def update_entry(self, glossary_id: str, data: GlossaryEntryUpdate) -> dict[str, Any]:
        """Apply a partial update to a glossary entry."""
        existing = self._repo.get_by_id(glossary_id)
        if not existing:
            raise HTTPException(status_code=404, detail="Glossary entry not found")

        fields = data.model_dump(exclude_unset=True)
        if "scope" in fields and fields["scope"] is not None:
            fields["scope"] = fields["scope"].value
        if "translations" in fields and fields["translations"] is not None:
            fields["translations"] = {
                k: v for k, v in fields["translations"].items() if v is not None
            }

        if not fields:
            return existing

        try:
            self._repo.update(glossary_id, **fields)
        except Exception as exc:
            if "UNIQUE constraint failed" in str(exc):
                raise HTTPException(
                    status_code=409,
                    detail="Another glossary entry with this term and scope already exists",
                )
            raise

        updated = self._repo.get_by_id(glossary_id)
        logger.info("Updated glossary entry: glossary_id=%s", glossary_id)
        return updated  # type: ignore[return-value]

    # ── Delete ──────────────────────────────────────────────────────

    def delete_entry(self, glossary_id: str) -> dict[str, str]:
        """Delete a glossary entry."""
        deleted = self._repo.delete(glossary_id)
        if not deleted:
            raise HTTPException(status_code=404, detail="Glossary entry not found")
        logger.info("Deleted glossary entry: glossary_id=%s", glossary_id)
        return {"detail": "Glossary entry deleted"}

    # ── Excel Import ────────────────────────────────────────────────

    def import_from_excel(
        self, file_bytes: bytes, scope: str = "global"
    ) -> dict[str, Any]:
        """
        Parse a client-format .xlsx file and upsert glossary entries.

        The client Excel has:
          - A data sheet named ``EN>XX terms`` (e.g. ``EN>DE terms``).
          - Row 1 is blank; row 2 contains headers:
            ``English | Comments | en->XX | XX comments``
          - Row 3+ is data.
          - "No translation" prefix in the translation cell means
            ``do_not_translate = True``.  A parenthetical hint like
            ``(Lerninhaltsentwickler)`` is stored as a language comment.

        Args:
            file_bytes: Raw bytes of the uploaded .xlsx file.
            scope: Scope to assign to every imported entry (default ``global``).

        Returns:
            Summary dict with ``imported``, ``updated``, ``skipped``, and
            ``errors`` keys.
        """
        scope = scope.lower()
        if scope not in _VALID_SCOPES:
            raise HTTPException(
                status_code=400,
                detail=f"Invalid scope '{scope}'. Must be one of: {', '.join(sorted(_VALID_SCOPES))}",
            )

        try:
            wb = load_workbook(filename=io.BytesIO(file_bytes), read_only=True)
        except Exception as exc:
            raise HTTPException(
                status_code=400,
                detail=f"Cannot read Excel file: {exc}",
            )

        # Detect the data sheet and language
        lang, ws = self._find_data_sheet(wb)

        # Read all rows; headers are at row 2 (index 1)
        rows = list(ws.iter_rows(values_only=True))
        if len(rows) < 2:
            wb.close()
            raise HTTPException(
                status_code=400,
                detail="Excel data sheet has fewer than 2 rows (header expected at row 2)",
            )

        header = [str(c).strip().lower() if c else "" for c in rows[1]]
        col_map = self._build_col_map(header, lang)

        imported = 0
        updated = 0
        skipped = 0
        errors: list[dict[str, Any]] = []

        for row_num, row in enumerate(rows[2:], start=3):
            try:
                term = self._cell_str(row, col_map.get("english"))
                if not term:
                    continue  # skip blank rows

                comments = self._cell_str(row, col_map.get("comments")) or None
                translation_raw = self._cell_str(row, col_map.get("translation"))
                lang_comments = self._cell_str(row, col_map.get("lang_comments")) or None

                # Parse "No translation" logic
                do_not_translate, translation, hint = self._parse_translation(translation_raw)

                # Merge parenthetical hint into lang_comments if present
                if hint and not lang_comments:
                    lang_comments = hint

                # Build translations dict
                translations: dict[str, str] | None = None
                if not do_not_translate and translation:
                    translations = {lang: translation}

                # Build per-language comment kwargs
                de_comments = lang_comments if lang == "de" else None
                es_comments = lang_comments if lang == "es" else None

                # Upsert: check if entry already exists
                existing = self._repo.get_by_term_and_scope(term, scope)

                if existing:
                    update_fields: dict[str, Any] = {
                        "do_not_translate": do_not_translate,
                    }
                    # Merge translations with existing ones (other language may exist)
                    existing_trans = existing.get("translations") or {}
                    if translations:
                        existing_trans.update(translations)
                    elif do_not_translate:
                        existing_trans.pop(lang, None)
                    if existing_trans:
                        update_fields["translations"] = existing_trans
                    else:
                        update_fields["translations"] = None
                    if comments is not None:
                        update_fields["comments"] = comments
                    if de_comments is not None:
                        update_fields["de_comments"] = de_comments
                    if es_comments is not None:
                        update_fields["es_comments"] = es_comments
                    self._repo.update(existing["glossary_id"], **update_fields)
                    updated += 1
                else:
                    self._repo.create(
                        glossary_id=uuid.uuid4().hex,
                        term=term,
                        scope=scope,
                        do_not_translate=do_not_translate,
                        translations=translations,
                        comments=comments,
                        de_comments=de_comments,
                        es_comments=es_comments,
                    )
                    imported += 1

            except Exception as exc:
                errors.append({"row": row_num, "error": str(exc)})

        wb.close()
        logger.info(
            "Glossary import complete: lang=%s, scope=%s, imported=%d, updated=%d, skipped=%d, errors=%d",
            lang, scope, imported, updated, skipped, len(errors),
        )
        return {
            "imported": imported,
            "updated": updated,
            "skipped": skipped,
            "language": lang,
            "scope": scope,
            "errors": errors,
        }

    # ── Excel Export ────────────────────────────────────────────────

    def export_to_excel(
        self, language: str, scope: Optional[str] = None
    ) -> bytes:
        """
        Build a client-format .xlsx from glossary entries for a single language.

        Args:
            language: Target language code (``de`` or ``es``).
            scope: Optional scope filter.
        """
        lang = language.lower()
        label = _LANG_LABELS[lang]
        items = self._repo.list_all(scope=scope)

        wb = Workbook()
        ws = wb.active
        ws.title = f"EN>{label} terms"

        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")

        headers = ["English", "Comments", f"en->{label}", f"{label} comments"]
        # Row 1 intentionally blank (client format)
        for col_idx, col_name in enumerate(headers, start=1):
            cell = ws.cell(row=2, column=col_idx, value=col_name)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        # Data rows start at row 3
        for row_idx, item in enumerate(items, start=3):
            translations = item.get("translations") or {}
            translation_val = translations.get(lang, "")
            lang_comment = item.get(f"{lang}_comments", "") or ""

            # Reconstruct "No translation (hint)" for do_not_translate entries
            if item.get("do_not_translate"):
                if lang_comment:
                    translation_val = f"No translation ({lang_comment})"
                    lang_comment = ""  # hint already embedded
                else:
                    translation_val = "No translation"

            ws.cell(row=row_idx, column=1, value=item.get("term", ""))
            ws.cell(row=row_idx, column=2, value=item.get("comments", "") or "")
            ws.cell(row=row_idx, column=3, value=translation_val)
            ws.cell(row=row_idx, column=4, value=lang_comment)

        # Auto-width
        for col_idx in range(1, 5):
            max_len = len(headers[col_idx - 1])
            for row in ws.iter_rows(min_row=3, min_col=col_idx, max_col=col_idx, values_only=True):
                val = str(row[0]) if row[0] else ""
                max_len = max(max_len, len(val))
            ws.column_dimensions[ws.cell(row=2, column=col_idx).column_letter].width = max_len + 4

        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    # ── Excel Template ──────────────────────────────────────────────

    def get_template_excel(self, language: str) -> bytes:
        """Return a blank client-format .xlsx template with example rows."""
        lang = language.lower()
        label = _LANG_LABELS[lang]

        wb = Workbook()
        ws = wb.active
        ws.title = f"EN>{label} terms"

        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")

        headers = ["English", "Comments", f"en->{label}", f"{label} comments"]
        for col_idx, col_name in enumerate(headers, start=1):
            cell = ws.cell(row=2, column=col_idx, value=col_name)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        # Example translations per language
        examples = {
            "de": [
                ["Active Ingredient", "Also known as API", "Wirkstoff", ""],
                ["Spiriva Respimat", "Brand name", "No translation", ""],
                ["Content Owner", "Neu aufgenommen", "No translation (Lerninhaltsentwickler)", ""],
            ],
            "es": [
                ["Active Ingredient", "Also known as API", "Principio activo", ""],
                ["Spiriva Respimat", "Brand name", "No translation", ""],
                ["Content Owner", "", "", ""],
            ],
        }

        for r_offset, row_data in enumerate(examples.get(lang, examples["de"]), start=3):
            for col_idx, val in enumerate(row_data, start=1):
                ws.cell(row=r_offset, column=col_idx, value=val)

        # Auto-width
        for col_idx in range(1, 5):
            max_len = len(headers[col_idx - 1])
            for row in ws.iter_rows(min_row=3, min_col=col_idx, max_col=col_idx, values_only=True):
                val = str(row[0]) if row[0] else ""
                max_len = max(max_len, len(val))
            ws.column_dimensions[ws.cell(row=2, column=col_idx).column_letter].width = max_len + 4

        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    # ── Helpers ─────────────────────────────────────────────────────

    @staticmethod
    def _cell_str(row: tuple, idx: int | None) -> str:
        """Safely extract a cell value as a stripped string."""
        if idx is None or idx >= len(row) or row[idx] is None:
            return ""
        return str(row[idx]).strip()

    @staticmethod
    def _find_data_sheet(wb: Any) -> tuple[str, Any]:
        """
        Scan workbook sheet names for the ``EN>XX terms`` pattern.

        Returns (language_code, worksheet). Raises 400 if not found.
        """
        for name in wb.sheetnames:
            m = _SHEET_PATTERN.match(name)
            if m:
                lang = m.group(1).lower()
                if lang not in _VALID_LANGUAGES:
                    raise HTTPException(
                        status_code=400,
                        detail=(
                            f"Unsupported language '{lang}' in sheet '{name}'."
                            f" Supported: {', '.join(sorted(_VALID_LANGUAGES))}"
                        ),
                    )
                return lang, wb[name]
        raise HTTPException(
            status_code=400,
            detail=(
                "No data sheet found. Expected a sheet named 'EN>DE terms' or "
                "'EN>ES terms'. "
                f"Found sheets: {', '.join(wb.sheetnames)}"
            ),
        )

    @staticmethod
    def _build_col_map(header: list[str], lang: str) -> dict[str, int]:
        """
        Map normalised header names to column indices.

        Expected headers (case-insensitive):
        ``English``, ``Comments``, ``en->XX``, ``XX comments``
        """
        col_map: dict[str, int] = {}
        label = lang.upper()

        for idx, h in enumerate(header):
            h_lower = h.lower().strip()
            if h_lower == "english":
                col_map["english"] = idx
            elif h_lower == "comments":
                col_map["comments"] = idx
            elif h_lower == f"en->{lang}" or h_lower == f"en->{label.lower()}":
                col_map["translation"] = idx
            elif h_lower == f"{lang} comments" or h_lower == f"{label.lower()} comments":
                col_map["lang_comments"] = idx

        if "english" not in col_map:
            raise HTTPException(
                status_code=400,
                detail=f"Missing required column 'English' in header row. Found: {header}",
            )
        return col_map

    @staticmethod
    def _parse_translation(raw: str) -> tuple[bool, str, str | None]:
        """
        Parse a translation cell value.

        Returns ``(do_not_translate, translation_text, parenthetical_hint)``.

        Examples:
          - ``"Wirkstoff"``                          → (False, "Wirkstoff", None)
          - ``"No translation"``                     → (True,  "", None)
          - ``"No translation (Lerninhaltsentwickler)"`` → (True,  "", "Lerninhaltsentwickler")
          - ``""``                                   → (False, "", None)
        """
        if not raw:
            return False, "", None

        if _NO_TRANSLATE_PREFIX.match(raw):
            hint_match = _PAREN_HINT.search(raw)
            hint = hint_match.group(1).strip() if hint_match else None
            return True, "", hint

        return False, raw, None
