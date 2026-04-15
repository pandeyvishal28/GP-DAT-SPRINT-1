"""
services/glossary_service.py
Business logic for glossary CRUD operations.
"""

from __future__ import annotations

import io
import logging
import uuid
from typing import Any, Optional

from fastapi import HTTPException
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill

from models.schemas import GlossaryEntryCreate, GlossaryEntryUpdate
from repositories.glossary_repository import GlossaryRepository

logger = logging.getLogger(__name__)

_EXCEL_COLUMNS = ["term", "scope", "do_not_translate", "de", "es", "notes"]
_VALID_SCOPES = {"global", "local", "functional"}
_TRUTHY = {"true", "yes", "1"}
_FALSY = {"false", "no", "0", ""}


class GlossaryService:
    """Encapsulates glossary business rules on top of the repository layer."""

    def __init__(self, repository: GlossaryRepository) -> None:
        self._repo = repository

    # ── Create ──────────────────────────────────────────────────────

    def create_entry(self, data: GlossaryEntryCreate) -> dict[str, Any]:
        """Create a new glossary entry. Raises 409 on duplicate term+scope."""
        entry_id = uuid.uuid4().hex

        try:
            self._repo.create(
                entry_id=entry_id,
                term=data.term,
                scope=data.scope.value,
                do_not_translate=data.do_not_translate,
                translations=(
                    data.translations.model_dump(exclude_none=True)
                    if data.translations else None
                ),
                notes=data.notes,
            )
        except Exception as exc:
            if "UNIQUE constraint failed" in str(exc):
                raise HTTPException(
                    status_code=409,
                    detail=f"Glossary entry '{data.term}' already exists for scope '{data.scope.value}'",
                )
            raise

        entry = self._repo.get_by_id(entry_id)
        logger.info("Created glossary entry: id=%s, term='%s'", entry_id, data.term)
        return entry  # type: ignore[return-value]

    # ── List ────────────────────────────────────────────────────────

    def list_entries(self, scope: Optional[str] = None) -> dict[str, Any]:
        """Return all active glossary entries, optionally filtered by scope."""
        items = self._repo.list_all(scope=scope)
        return {"items": items, "total": len(items)}

    # ── Get ─────────────────────────────────────────────────────────

    def get_entry(self, entry_id: str) -> dict[str, Any]:
        """Fetch a single glossary entry or raise 404."""
        entry = self._repo.get_by_id(entry_id)
        if not entry:
            raise HTTPException(status_code=404, detail="Glossary entry not found")
        return entry

    # ── Update ──────────────────────────────────────────────────────

    def update_entry(self, entry_id: str, data: GlossaryEntryUpdate) -> dict[str, Any]:
        """Apply a partial update to a glossary entry."""
        existing = self._repo.get_by_id(entry_id)
        if not existing:
            raise HTTPException(status_code=404, detail="Glossary entry not found")

        fields = data.model_dump(exclude_unset=True)
        if "scope" in fields and fields["scope"] is not None:
            fields["scope"] = fields["scope"].value
        if "translations" in fields and fields["translations"] is not None:
            fields["translations"] = {
                k: v for k, v in fields["translations"].items() if v is not None
            }

        if not fields:
            return existing

        try:
            self._repo.update(entry_id, **fields)
        except Exception as exc:
            if "UNIQUE constraint failed" in str(exc):
                raise HTTPException(
                    status_code=409,
                    detail="Another glossary entry with this term and scope already exists",
                )
            raise

        updated = self._repo.get_by_id(entry_id)
        logger.info("Updated glossary entry: id=%s", entry_id)
        return updated  # type: ignore[return-value]

    # ── Delete ──────────────────────────────────────────────────────

    def delete_entry(self, entry_id: str) -> dict[str, str]:
        """Delete a glossary entry."""
        deleted = self._repo.delete(entry_id)
        if not deleted:
            raise HTTPException(status_code=404, detail="Glossary entry not found")
        logger.info("Deleted glossary entry: id=%s", entry_id)
        return {"detail": "Glossary entry deleted"}

    # ── Excel Import ────────────────────────────────────────────────

    def import_from_excel(self, file_bytes: bytes) -> dict[str, Any]:
        """
        Parse an .xlsx file and upsert glossary entries.

        Returns a summary with counts of imported, updated, and errored rows.
        """
        try:
            wb = load_workbook(filename=io.BytesIO(file_bytes), read_only=True)
        except Exception as exc:
            raise HTTPException(
                status_code=400,
                detail=f"Cannot read Excel file: {exc}",
            )

        ws = wb.active
        if ws is None:
            raise HTTPException(status_code=400, detail="Excel file has no active sheet")

        # Read header row
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            raise HTTPException(status_code=400, detail="Excel file is empty")

        header = [str(c).strip().lower() if c else "" for c in rows[0]]

        # Validate required columns exist
        missing = [c for c in ("term", "scope", "do_not_translate") if c not in header]
        if missing:
            raise HTTPException(
                status_code=400,
                detail=f"Missing required columns: {', '.join(missing)}",
            )

        col_map = {name: idx for idx, name in enumerate(header) if name in _EXCEL_COLUMNS}

        imported = 0
        updated = 0
        errors: list[dict[str, Any]] = []

        for row_num, row in enumerate(rows[1:], start=2):
            try:
                term = self._cell_str(row, col_map.get("term"))
                scope = self._cell_str(row, col_map.get("scope")).lower()
                dnt_raw = self._cell_str(row, col_map.get("do_not_translate")).lower()
                de = self._cell_str(row, col_map.get("de"))
                es = self._cell_str(row, col_map.get("es"))
                notes = self._cell_str(row, col_map.get("notes")) or None

                # Skip completely empty rows
                if not term:
                    continue

                # Validate scope
                if scope not in _VALID_SCOPES:
                    errors.append({"row": row_num, "error": f"Invalid scope '{scope}'. Must be one of: global, local, functional"})
                    continue

                # Parse do_not_translate
                if dnt_raw in _TRUTHY:
                    do_not_translate = True
                elif dnt_raw in _FALSY:
                    do_not_translate = False
                else:
                    errors.append({"row": row_num, "error": f"Invalid do_not_translate value '{dnt_raw}'. Use true/false, yes/no, or 1/0"})
                    continue

                # Validate translations
                if not do_not_translate and not de and not es:
                    errors.append({"row": row_num, "error": "At least one translation (de or es) is required when do_not_translate is false"})
                    continue

                translations = {}
                if de:
                    translations["de"] = de
                if es:
                    translations["es"] = es

                # Upsert: check if entry already exists
                existing = self._repo.get_by_term_and_scope(term, scope)

                if existing:
                    self._repo.update(
                        existing["id"],
                        do_not_translate=do_not_translate,
                        translations=translations if translations else None,
                        notes=notes,
                    )
                    updated += 1
                else:
                    self._repo.create(
                        entry_id=uuid.uuid4().hex,
                        term=term,
                        scope=scope,
                        do_not_translate=do_not_translate,
                        translations=translations if translations else None,
                        notes=notes,
                    )
                    imported += 1

            except Exception as exc:
                errors.append({"row": row_num, "error": str(exc)})

        wb.close()
        logger.info(
            "Glossary import complete: imported=%d, updated=%d, errors=%d",
            imported, updated, len(errors),
        )
        return {"imported": imported, "updated": updated, "errors": errors}

    # ── Excel Export ────────────────────────────────────────────────

    def export_to_excel(self, scope: Optional[str] = None) -> bytes:
        """Build an .xlsx workbook from glossary entries and return as bytes."""
        items = self._repo.list_all(scope=scope)

        wb = Workbook()
        ws = wb.active
        ws.title = "Glossary"

        # Header style
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")

        for col_idx, col_name in enumerate(_EXCEL_COLUMNS, start=1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        # Data rows
        for row_idx, item in enumerate(items, start=2):
            translations = item.get("translations") or {}
            ws.cell(row=row_idx, column=1, value=item.get("term", ""))
            ws.cell(row=row_idx, column=2, value=item.get("scope", ""))
            ws.cell(row=row_idx, column=3, value=str(item.get("do_not_translate", False)).lower())
            ws.cell(row=row_idx, column=4, value=translations.get("de", ""))
            ws.cell(row=row_idx, column=5, value=translations.get("es", ""))
            ws.cell(row=row_idx, column=6, value=item.get("notes", "") or "")

        # Auto-width columns
        for col_idx, col_name in enumerate(_EXCEL_COLUMNS, start=1):
            max_len = len(col_name)
            for row in ws.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx, values_only=True):
                val = str(row[0]) if row[0] else ""
                max_len = max(max_len, len(val))
            ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = max_len + 4

        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    # ── Excel Template ──────────────────────────────────────────────

    def get_template_excel(self) -> bytes:
        """Return a blank .xlsx template with headers and one example row."""
        wb = Workbook()
        ws = wb.active
        ws.title = "Glossary"

        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")

        for col_idx, col_name in enumerate(_EXCEL_COLUMNS, start=1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        # Example row
        example = [
            "Active Ingredient",
            "global",
            "false",
            "Wirkstoff",
            "Principio activo",
            "Also known as API",
        ]
        for col_idx, val in enumerate(example, start=1):
            ws.cell(row=2, column=col_idx, value=val)

        # Second example — keep as-is
        example2 = ["Spiriva Respimat", "global", "true", "", "", "Brand name — keep as-is"]
        for col_idx, val in enumerate(example2, start=1):
            ws.cell(row=3, column=col_idx, value=val)

        # Auto-width
        for col_idx, col_name in enumerate(_EXCEL_COLUMNS, start=1):
            max_len = max(len(col_name), len(example[col_idx - 1]), len(example2[col_idx - 1]))
            ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = max_len + 4

        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    # ── Helpers ─────────────────────────────────────────────────────

    @staticmethod
    def _cell_str(row: tuple, idx: int | None) -> str:
        """Safely extract a cell value as a stripped string."""
        if idx is None or idx >= len(row) or row[idx] is None:
            return ""
        return str(row[idx]).strip()
